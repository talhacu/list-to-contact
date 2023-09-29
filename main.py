import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, Label, Entry, Button, PhotoImage
import time


def convert_excel_to_vcf():
    input_file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

    if input_file:
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        rows_to_delete = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            if all([cell.value is None for cell in row]):
                rows_to_delete.append(row)

        for row in rows_to_delete:
            sheet.delete_rows(row[0].row)
        workbook.save(input_file)

        output_file = f"{input_file}_KİŞİLER.vcf"
        baslik = kisilerin_basi_entry.get()
        son = kisilerin_sonu_entry.get()

        with open(output_file, "w", encoding="utf-8") as vcf_file:
            for row in sheet.iter_rows(values_only=True):
                name = row[0]
                phone = row[1]
                phone = str(phone).replace(" ", "").zfill(11)
                full_name = f"{baslik} {name} {son}"

                vcf_file.write("BEGIN:VCARD\n")
                vcf_file.write("VERSION:3.0\n")
                vcf_file.write(f"N:{full_name}\n")
                vcf_file.write(f"TEL:{phone}\n")
                vcf_file.write("END:VCARD\n")

        time.sleep(1.5)
        messagebox.showinfo(title="Tamamlandı", message="Liste başarıyla kişiler dosyasına dönüştürüldü")


root = tk.Tk()
root.title("Listeden Kişiler Dosyası Oluşturma")
root.geometry("500x500")

kisilerin_basi = Label(root, text="İsimlerin başında ne yazacak? [Örnek_Metin] Talha Bakır")
kisilerin_basi.grid(row=0, column=0, sticky="w", padx=20, pady=20)
kisilerin_basi_entry = Entry(root)
kisilerin_basi_entry.grid(row=0, column=1, padx=20, pady=20)

kisilerin_sonu = Label(root, text="İsimlerin sonunda ne yazacak? Talha Bakır [Örnek_Metin]")
kisilerin_sonu.grid(row=1, column=0, sticky="w", padx=20, pady=20)
kisilerin_sonu_entry = Entry(root)
kisilerin_sonu_entry.grid(row=1, column=1, padx=20, pady=20)

image = PhotoImage(file="liste.png")
resized = image.subsample(3)
label = Label(root, image=resized)
label.grid(row=2, columnspan=2, padx=20, pady=20)

aciklama = Label(root, text="Listenizi yukarıdaki şekilde düzenlemelisiniz \n A Sütununda İsim ve Soyisim \n B Sütununda ise başında sıfır olmadan numaralar olmalı")
aciklama.grid(row=3, columnspan=2, padx=20, pady=20)

convert_button = Button(root, text="Liste Dosyasını Seç ve Dönüştür", command=convert_excel_to_vcf)
convert_button.grid(row=4, columnspan=2, padx=20, pady=20)

root.mainloop()
